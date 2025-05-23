from flask import Flask, render_template, request, send_file
import pandas as pd
import re
import os
import numpy as np
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from werkzeug.utils import secure_filename
import gc
from functools import lru_cache

app = Flask(__name__)

# Configuration
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024  # 20 MB limit

# Compile regex patterns once at module level
PATTERN_NEFT = re.compile(r"(NEFT-|MClick/To\s+)")
PATTERN_PARENS = re.compile(r'\((.*?)\)?$')
PATTERN_NUMERIC_CLEAN = re.compile(r'[\$,\s]+')
PATTERN_UPI = re.compile(r'UPI:', re.IGNORECASE)
PATTERN_BALANCE_CLEAN = re.compile(r'[,\s]+')

def save_upload(file):
    """Save uploaded file and return path"""
    path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
    file.save(path)
    return path

@lru_cache(maxsize=128)
def get_engine_type(filename):
    """Cached engine type determination"""
    return 'openpyxl' if filename.endswith('.xlsx') else 'xlrd'

def load_account_statement(file_path):
    """Optimized account statement loading"""
    engine = get_engine_type(file_path)
    
    # Use chunking for large files and specify dtypes
    try:
        df = pd.read_excel(
            file_path, 
            header=None, 
            engine=engine,
            skiprows=14,  # Skip first 14 rows directly
            dtype=str,    # Read as string to avoid type inference overhead
            na_filter=False  # Don't convert to NaN, keep as strings
        )
    except Exception:
        # Fallback for older format
        df = pd.read_excel(file_path, header=None, engine=engine)
        df = df.drop(index=range(14))
    
    # Remove empty columns efficiently
    df = df.loc[:, df.notna().any()]
    
    # Set headers and clean
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    
    # Drop cheque column if exists
    if 'Cheque No' in df.columns:
        df = df.drop('Cheque No', axis=1)
    
    df.columns = ['Date', 'Particular', 'Given', 'Received', 'Balance']
    
    return df

def load_our_books(file_path):
    """Optimized our books loading"""
    engine = get_engine_type(file_path)
    
    df = pd.read_excel(
        file_path, 
        header=None, 
        engine=engine,
        dtype=str,
        na_filter=False
    )
    
    # Remove empty columns
    df = df.loc[:, df.notna().any()]
    
    # Set headers
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    
    # Filter out opening/closing balance rows
    mask = ~df['Particular'].str.contains(
        'Opening balance|Closing balance', 
        case=False, 
        na=False
    )
    df = df[mask].reset_index(drop=True)
    
    return df

def clean_numeric_column(series):
    """Efficiently clean numeric columns"""
    # Use vectorized operations
    cleaned = series.str.replace(PATTERN_NUMERIC_CLEAN, '', regex=True)
    cleaned = cleaned.replace(['', '0', 'None'], np.nan)
    return pd.to_numeric(cleaned, errors='coerce')

def parse_dates_efficiently(series):
    """Efficiently parse dates"""
    return pd.to_datetime(series, errors='coerce', dayfirst=True, cache=True).dt.date

def process_receivements(ac_path, books_path, timestamp):
    """Optimized receivements processing"""
    # Load data
    ac = load_account_statement(ac_path)
    ob = load_our_books(books_path)
    
    # Filter receivements early to reduce memory usage
    recv_ac = ac[ac['Received'].notna() & (ac['Received'] != '')].copy()
    recv_ob = ob[ob['Debit'].notna() & (ob['Debit'] != '') & (ob['Debit'] != '0')].copy()
    
    # Early cleanup
    del ac, ob
    gc.collect()
    
    if recv_ac.empty or recv_ob.empty:
        return pd.DataFrame(columns=['Date', 'Not in account statement', 'Debit', '', 'Not in our books', 'Received'])
    
    # Clean supplier names efficiently
    recv_ac['SupplierName'] = recv_ac['Particular'].str.strip()
    recv_ob['Particular'] = recv_ob['Particular'].str.strip()
    
    # Extract text inside parentheses for our books
    matches = recv_ob['Particular'].str.extract(PATTERN_PARENS, expand=False)
    recv_ob['Particular'] = matches.fillna(recv_ob['Particular'])
    
    # Parse dates
    recv_ac['Date'] = parse_dates_efficiently(recv_ac['Date'])
    recv_ob['Date'] = parse_dates_efficiently(recv_ob['Date'])
    
    # Clean numeric columns
    recv_ac['Received'] = clean_numeric_column(recv_ac['Received'])
    recv_ob['Debit'] = clean_numeric_column(recv_ob['Debit'])
    
    # Get unique dates
    dates = sorted(set(recv_ac['Date'].dropna()) | set(recv_ob['Date'].dropna()))
    
    # Process discrepancies efficiently
    discrepancies = []
    
    for date in dates:
        ac_day = recv_ac[recv_ac['Date'] == date]
        ob_day = recv_ob[recv_ob['Date'] == date]
        
        names_ac = set(ac_day['SupplierName'].dropna())
        names_ob = set(ob_day['Particular'].dropna())
        
        only_ob = names_ob - names_ac
        only_ac = names_ac - names_ob
        
        # Create balanced lists
        max_len = max(len(only_ob), len(only_ac), 1)
        only_ob_list = list(only_ob) + [None] * (max_len - len(only_ob))
        only_ac_list = list(only_ac) + [None] * (max_len - len(only_ac))
        
        for ob_name, ac_name in zip(only_ob_list, only_ac_list):
            discrepancies.append({
                'Date': date,
                'Not in account statement': ob_name,
                'Not in our books': ac_name
            })
    
    if not discrepancies:
        return pd.DataFrame(columns=['Date', 'Not in account statement', 'Debit', '', 'Not in our books', 'Received'])
    
    df = pd.DataFrame(discrepancies)
    
    # Calculate sums efficiently
    sum_ac = recv_ac.groupby(['Date', 'SupplierName'], as_index=False)['Received'].sum()
    sum_ob = recv_ob.groupby(['Date', 'Particular'], as_index=False)['Debit'].sum()
    
    # Merge sums
    df = df.merge(
        sum_ac, 
        left_on=['Date', 'Not in our books'], 
        right_on=['Date', 'SupplierName'], 
        how='left'
    ).drop(columns=['SupplierName'])
    
    df = df.merge(
        sum_ob,
        left_on=['Date', 'Not in account statement'],
        right_on=['Date', 'Particular'],
        how='left'
    ).drop(columns=['Particular'])
    
    # Sort and add blank rows
    df = df.sort_values('Date').reset_index(drop=True)
    
    # Add blank rows between dates efficiently
    rows = []
    prev_date = None
    
    for _, row in df.iterrows():
        if prev_date is not None and row['Date'] != prev_date:
            rows.append({col: None for col in df.columns})
        rows.append(row.to_dict())
        prev_date = row['Date']
    
    final = pd.DataFrame(rows)
    final[''] = None
    final = final[['Date', 'Not in account statement', 'Debit', '', 'Not in our books', 'Received']]
    
    # Filter UPI transactions
    upi_mask = (
        ~final['Not in account statement'].str.contains(PATTERN_UPI, na=False) &
        ~final['Not in our books'].str.contains(PATTERN_UPI, na=False)
    )
    final = final[upi_mask]
    
    return final

def process_payments(ac_path, books_path, timestamp):
    """Optimized payments processing"""
    ac = load_account_statement(ac_path)
    ob = load_our_books(books_path)
    
    # Filter payments early
    pay_ac = ac[ac['Given'].notna() & (ac['Given'] != '')].copy()
    pay_ob = ob[ob['Credit'].notna() & (ob['Credit'] != '') & (ob['Credit'] != '0')].copy()
    
    # Early cleanup
    del ac, ob
    gc.collect()
    
    if pay_ac.empty or pay_ob.empty:
        return pd.DataFrame(columns=['Date', 'Not in account statement', 'Credit', '', 'Not in our books', 'Given'])
    
    pay_ob['Particular'] = pay_ob['Particular'].str.strip()
    
    # Clean particular names efficiently
    pay_ac['Particular'] = pay_ac['Particular'].str.replace(PATTERN_NEFT, '', regex=True).str.strip()
    pay_ac['SupplierName'] = pay_ac['Particular'].str.split('/').str[0].str.strip()
    
    # Parse dates
    pay_ac['Date'] = parse_dates_efficiently(pay_ac['Date'])
    pay_ob['Date'] = parse_dates_efficiently(pay_ob['Date'])
    
    # Clean numeric columns
    pay_ac['Given'] = clean_numeric_column(pay_ac['Given'])
    pay_ob['Credit'] = clean_numeric_column(pay_ob['Credit'])
    
    # Get unique dates
    dates = sorted(set(pay_ac['Date'].dropna()) | set(pay_ob['Date'].dropna()))
    
    # Process discrepancies
    discrepancies = []
    
    for date in dates:
        ac_day = pay_ac[pay_ac['Date'] == date]
        ob_day = pay_ob[pay_ob['Date'] == date]
        
        names_ac = set(ac_day['SupplierName'].dropna())
        names_ob = set(ob_day['Particular'].dropna())
        
        only_ob = list(names_ob)
        only_ac = list(names_ac)
        
        max_len = max(len(only_ob), len(only_ac), 1)
        only_ob += [None] * (max_len - len(only_ob))
        only_ac += [None] * (max_len - len(only_ac))
        
        for ob_name, ac_name in zip(only_ob, only_ac):
            discrepancies.append({
                'Date': date,
                'Not in account statement': ob_name,
                'Not in our books': ac_name
            })
    
    if not discrepancies:
        return pd.DataFrame(columns=['Date', 'Not in account statement', 'Credit', '', 'Not in our books', 'Given'])
    
    df = pd.DataFrame(discrepancies)
    
    # Calculate sums
    sum_ac = pay_ac.groupby(['Date', 'SupplierName'], as_index=False)['Given'].sum()
    sum_ob = pay_ob.groupby(['Date', 'Particular'], as_index=False)['Credit'].sum()
    
    # Merge sums
    df = df.merge(
        sum_ac,
        left_on=['Date', 'Not in our books'],
        right_on=['Date', 'SupplierName'],
        how='left'
    ).drop(columns=['SupplierName'])
    
    df = df.merge(
        sum_ob,
        left_on=['Date', 'Not in account statement'],
        right_on=['Date', 'Particular'],
        how='left'
    ).drop(columns=['Particular'])
    
    # Sort and add blank rows
    df = df.sort_values('Date').reset_index(drop=True)
    
    rows = []
    prev_date = None
    
    for _, row in df.iterrows():
        if prev_date is not None and row['Date'] != prev_date:
            rows.append({col: None for col in df.columns})
        rows.append(row.to_dict())
        prev_date = row['Date']
    
    final = pd.DataFrame(rows)
    final[''] = None
    final = final[['Date', 'Not in account statement', 'Credit', '', 'Not in our books', 'Given']]
    
    return final

def process_summary(ac_path, books_path):
    """Optimized summary processing"""
    ac = load_account_statement(ac_path)
    ob = load_our_books(books_path)
    
    # Parse dates efficiently
    ac['Date'] = pd.to_datetime(ac['Date'], errors='coerce', dayfirst=True, cache=True)
    ob['Date'] = pd.to_datetime(ob['Date'], errors='coerce', dayfirst=True, cache=True)
    
    # Clean numeric columns efficiently
    numeric_cols_ac = ['Given', 'Received']
    numeric_cols_ob = ['Credit', 'Debit']
    
    for col in numeric_cols_ac:
        ac[col] = clean_numeric_column(ac[col])
    
    for col in numeric_cols_ob:
        ob[col] = clean_numeric_column(ob[col])
    
    # Clean balance columns
    ac['Balance'] = (
        ac['Balance'].astype(str)
        .str.replace(PATTERN_BALANCE_CLEAN, '', regex=True)
        .pipe(pd.to_numeric, errors='coerce')
        .abs()
    )
    
    ob['Balance'] = (
        ob['Balance'].astype(str)
        .str.replace('Cr', '', regex=False)
        .str.replace(PATTERN_BALANCE_CLEAN, '', regex=True)
        .pipe(pd.to_numeric, errors='coerce')
        .abs()
    )
    
    # Get unique dates efficiently
    all_dates = sorted(
        set(ac['Date'].dropna().dt.date) | 
        set(ob['Date'].dropna().dt.date)
    )
    
    # Pre-calculate aggregations by date for efficiency
    ac_by_date = {}
    ob_by_date = {}
    
    for date in all_dates:
        ac_day = ac[ac['Date'].dt.date == date]
        ob_day = ob[ob['Date'].dt.date == date]
        
        ac_by_date[date] = {
            'count': len(ac_day),
            'received_count': ac_day['Received'].count(),
            'given_count': ac_day['Given'].count(),
            'received_sum': ac_day['Received'].sum(),
            'given_sum': ac_day['Given'].sum(),
            'balance': ac_day['Balance'].iloc[0] if not ac_day.empty else None
        }
        
        ob_by_date[date] = {
            'count': len(ob_day),
            'debit_count': ob_day['Debit'].count(),
            'credit_count': ob_day['Credit'].count(),
            'debit_sum': ob_day['Debit'].sum(),
            'credit_sum': ob_day['Credit'].sum(),
            'balance': ob_day['Balance'].iloc[-1] if not ob_day.empty else None
        }
    
    # Build summary efficiently
    rows = []
    for date in all_dates:
        ac_data = ac_by_date[date]
        ob_data = ob_by_date[date]
        
        metrics = [
            ('total count', ac_data['count'], ob_data['count']),
            ('debit/received entries', ac_data['received_count'], ob_data['debit_count']),
            ('credit/given entries', ac_data['given_count'], ob_data['credit_count']),
            ('debit/received total', ac_data['received_sum'], ob_data['debit_sum']),
            ('credit/given total', ac_data['given_sum'], ob_data['credit_sum']),
            ('Closing Balance', ac_data['balance'], ob_data['balance'])
        ]
        
        for metric, account_val, our_book_val in metrics:
            rows.append({
                'Date': date,
                'Metric': metric,
                'account': account_val,
                'our_book': our_book_val,
                'Difference': (account_val or 0) - (our_book_val or 0)
            })
    
    df = pd.DataFrame(rows)
    df = df.set_index(['Date', 'Metric'])
    
    return df

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process_files():
    if 'ac_statement' not in request.files or 'our_books' not in request.files:
        return "Error: Both files are required!", 400
    
    ac_file = request.files['ac_statement']
    ob_file = request.files['our_books']
    
    if ac_file.filename == '' or ob_file.filename == '':
        return "Error: Please select both files!", 400
    
    try:
        # Save uploaded files
        ac_path = save_upload(ac_file)
        ob_path = save_upload(ob_file)
        
        timestamp = int(time.time())
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], f"combined_{timestamp}.xlsx")
        
        # Process data
        recv_df = process_receivements(ac_path, ob_path, timestamp)
        pay_df = process_payments(ac_path, ob_path, timestamp)
        summary_df = process_summary(ac_path, ob_path)
        
        # Write to Excel efficiently
        with pd.ExcelWriter(output_path, engine='openpyxl', options={'remove_timezone': True}) as writer:
            recv_df.to_excel(writer, sheet_name='Receivements', index=False)
            pay_df.to_excel(writer, sheet_name='Payments', index=False)
            summary_df.to_excel(writer, sheet_name='Summary', index=True)
        
        # Apply styling
        wb = load_workbook(output_path)
        grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        for ws in wb.worksheets:
            if ws.title in ['Receivements', 'Payments']:
                for row_idx in range(2, ws.max_row + 1):
                    if ws.cell(row=row_idx, column=1).value is None:
                        for col_idx in range(1, min(7, ws.max_column + 1)):
                            ws.cell(row=row_idx, column=col_idx).fill = grey_fill
        
        wb.save(output_path)
        wb.close()
        
        # Clean up uploaded files
        try:
            os.remove(ac_path)
            os.remove(ob_path)
        except OSError:
            pass  # Files might be locked, ignore cleanup errors
        
        # Force garbage collection
        gc.collect()
        
        return send_file(output_path, as_attachment=True)
        
    except Exception as e:
        # Clean up on error
        try:
            if 'ac_path' in locals():
                os.remove(ac_path)
            if 'ob_path' in locals():
                os.remove(ob_path)
            if 'output_path' in locals() and os.path.exists(output_path):
                os.remove(output_path)
        except OSError:
            pass
        
        return f"Error processing files: {str(e)}", 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.getenv('PORT', 5000)), debug=False)